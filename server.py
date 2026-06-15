#!/usr/bin/env python3
"""Multi-project test report server (standalone QA tool).

เครื่องมือนี้ "อ่าน" ผลทดสอบจาก checkout ของ main repo (realfact) ตัวมันเองไม่ได้
เก็บ data — ฉะนั้นต้องบอกให้รู้ว่า main repo อยู่ที่ไหนผ่าน (เรียงตามลำดับความสำคัญ):
    1. --repo-root /path/to/realfact
    2. env REALFACT_REPO_ROOT=/path/to/realfact
    3. auto-detect: เดินขึ้นจาก cwd หา .git ที่มีโฟลเดอร์ tests/kiosk-avatar
    4. fallback: ../realfact ข้างๆ tool repo นี้

รัน:
    python3 server.py --repo-root /path/to/realfact [--port 8088]
    # หรือ
    REALFACT_REPO_ROOT=/path/to/realfact python3 server.py

เปิด browser ที่ http://localhost:8088
"""

import argparse
import base64
import csv
import glob
import json
import os
import re
import subprocess
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
import uuid
from datetime import datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from urllib.parse import urlparse, parse_qs

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
UI_DIR = os.path.join(SCRIPT_DIR, "ui")
CONFIG_PATH = os.path.join(SCRIPT_DIR, "projects.json")

# REPO_ROOT = path ของ main repo (realfact) ที่เก็บ test data จริง
# ตัว tool นี้เป็น standalone repo แยกออกมา จึงต้อง resolve ค่านี้ตอน startup
# (ดู resolve_repo_root() + main()) ก่อนเรียก load_projects()/load_jira_config()
REPO_ROOT = None
JIRA_CONFIG_PATH = None  # = REPO_ROOT/.claude/jira_config.md, ตั้งหลัง resolve REPO_ROOT


def resolve_repo_root(cli_value=None):
    """หา path ของ main repo (realfact) ตามลำดับความสำคัญ:
    1. --repo-root (cli_value)
    2. env REALFACT_REPO_ROOT
    3. auto-detect: เดินขึ้นจาก cwd หา dir ที่มี tests/kiosk-avatar
    4. fallback: ../realfact ข้างๆ tool repo นี้
    คืน absolute path (ยังไม่การันตีว่ามีจริง — ผู้เรียกควร validate)
    """
    candidate = cli_value or os.environ.get("REALFACT_REPO_ROOT")
    if candidate:
        return os.path.abspath(os.path.expanduser(candidate))

    # auto-detect จาก cwd เดินขึ้นไปเรื่อยๆ
    cur = os.path.abspath(os.getcwd())
    while True:
        if os.path.isdir(os.path.join(cur, "tests", "kiosk-avatar")):
            return cur
        parent = os.path.dirname(cur)
        if parent == cur:
            break
        cur = parent

    # fallback: sibling ../realfact
    return os.path.abspath(os.path.join(SCRIPT_DIR, "..", "realfact"))

ID_SAFE_RE = re.compile(r"^[A-Za-z0-9_.-]+$")

# JIRA integration — loaded once at startup
JIRA_CONFIG = None  # dict with base_url, auth_header, project_key
JIRA_CACHE = {}  # key: collection_id → {"ts": float, "issues": [...]}
JIRA_CACHE_TTL = 60  # seconds

# Run files cache — invalidated by mtime, so new runs always pick up
# key: file path → {"mtime": float, "data": dict}
RUN_FILE_CACHE = {}
# key: project["id"] → {"snapshot": list[(path, mtime)], "ts": float, "results": list[dict]}
LIST_RUNS_CACHE = {}
LIST_RUNS_TTL = 2.0  # seconds — short, just to dedup rapid reloads


def load_jira_config():
    """Read .claude/jira_config.md — returns dict or None if not present."""
    if not os.path.isfile(JIRA_CONFIG_PATH):
        return None
    try:
        text = open(JIRA_CONFIG_PATH, encoding="utf-8").read()
        base_url = re.search(r"JIRA_BASE_URL=(\S+)", text).group(1).rstrip("/")
        email = re.search(r"JIRA_EMAIL=(\S+)", text).group(1)
        token = re.search(r"JIRA_API_TOKEN=(\S+)", text).group(1)
        key = re.search(r"JIRA_PROJECT_KEY=(\S+)", text).group(1)
        auth = base64.b64encode(f"{email}:{token}".encode()).decode()
        return {
            "base_url": base_url,
            "auth_header": f"Basic {auth}",
            "project_key": key,
        }
    except Exception as e:
        print(f"[jira] config load failed: {e}")
        return None


def fetch_jira_issues_for_collection(cid, fresh=False):
    """Query JIRA for issues mentioning this collection_id.

    Uses JQL text search. Cached briefly to avoid spam on rapid reloads.
    Pass fresh=True to bypass cache (for user-triggered refresh).
    Returns list of dicts with {key, summary, type, status, assignee, url, created, labels}.
    """
    if not JIRA_CONFIG:
        return None  # not configured
    now = time.time()
    cached = JIRA_CACHE.get(cid)
    if not fresh and cached and (now - cached["ts"]) < JIRA_CACHE_TTL:
        return cached["issues"]

    jql = f'project = {JIRA_CONFIG["project_key"]} AND (labels = "{cid}" OR text ~ "{cid}")'
    qs = urllib.parse.urlencode({
        "jql": jql,
        "fields": "summary,status,assignee,issuetype,created,labels",
        "maxResults": 100,
    })
    url = f"{JIRA_CONFIG['base_url']}/rest/api/3/search/jql?{qs}"
    req = urllib.request.Request(url, headers={
        "Authorization": JIRA_CONFIG["auth_header"],
        "Accept": "application/json",
    })
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read().decode())
    except urllib.error.HTTPError as e:
        print(f"[jira] HTTP {e.code} when querying: {e.read().decode()[:200]}")
        return []
    except Exception as e:
        print(f"[jira] query failed: {e}")
        return []

    issues = []
    for i in data.get("issues", []):
        f = i.get("fields", {})
        a = f.get("assignee") or {}
        issues.append({
            "key": i["key"],
            "summary": f.get("summary", ""),
            "type": (f.get("issuetype") or {}).get("name", ""),
            "status": (f.get("status") or {}).get("name", ""),
            "status_category": ((f.get("status") or {}).get("statusCategory") or {}).get("key", ""),
            "assignee": a.get("displayName") or "Unassigned",
            "url": f"{JIRA_CONFIG['base_url']}/browse/{i['key']}",
            "created": f.get("created", ""),
            "labels": f.get("labels", []),
        })
    issues.sort(key=lambda x: x["key"])
    JIRA_CACHE[cid] = {"ts": now, "issues": issues}
    return issues


def load_projects():
    with open(CONFIG_PATH, encoding="utf-8") as f:
        cfg = json.load(f)
    for p in cfg["projects"]:
        p["_results_abs"] = os.path.join(REPO_ROOT, p["results_dir"])
        p["_collections_abs"] = os.path.join(REPO_ROOT, p["collections_dir"])
        os.makedirs(p["_collections_abs"], exist_ok=True)
    return cfg["projects"]


def find_project(pid, projects):
    return next((p for p in projects if p["id"] == pid), None)


def _load_run_file_cached(fp):
    """Read a result JSON, cached by mtime. Returns parsed dict or None."""
    try:
        mtime = os.path.getmtime(fp)
    except OSError:
        return None
    cached = RUN_FILE_CACHE.get(fp)
    if cached and cached["mtime"] == mtime:
        return cached["data"]
    try:
        with open(fp, encoding="utf-8") as f:
            d = json.load(f)
    except Exception:
        return None
    RUN_FILE_CACHE[fp] = {"mtime": mtime, "data": d}
    return d


def list_runs(project):
    """Return all run summaries (dedup by run_id, keep latest timestamp).

    Cached: skips re-glob + re-parse when no result file mtime changed.
    New runs invalidate automatically because their mtime is fresh and the
    file list snapshot differs.
    """
    results_dir = project["_results_abs"]
    if not os.path.isdir(results_dir):
        return []
    # Skip hidden directories (e.g. .backup_*) when walking
    files = []
    for root, dirs, fns in os.walk(results_dir):
        dirs[:] = [d for d in dirs if not d.startswith(".")]
        for fn in fns:
            if fn.startswith("result_") and fn.endswith(".json"):
                files.append(os.path.join(root, fn))

    # Build mtime snapshot — used as cache key
    try:
        snapshot = tuple(sorted((fp, os.path.getmtime(fp)) for fp in files))
    except OSError:
        snapshot = None

    pid = project.get("id", "")
    cached = LIST_RUNS_CACHE.get(pid)
    now = time.time()
    if cached and snapshot is not None and cached.get("snapshot") == snapshot \
            and (now - cached.get("ts", 0)) < LIST_RUNS_TTL:
        return cached["results"]

    by_run = {}
    for fp in files:
        d = _load_run_file_cached(fp)
        if d is None:
            continue
        rid = d.get("run_id", "")
        if not rid:
            continue
        existing = by_run.get(rid)
        if not existing or d.get("timestamp", "") > existing["_data"].get("timestamp", ""):
            by_run[rid] = {"_file": fp, "_data": d}

    results = list(by_run.values())
    if snapshot is not None:
        LIST_RUNS_CACHE[pid] = {"snapshot": snapshot, "ts": now, "results": results}
    return results


def load_run(project, run_id):
    for r in list_runs(project):
        if r["_data"].get("run_id") == run_id:
            return r
    return None


def list_collections(project):
    out = []
    for fp in sorted(glob.glob(os.path.join(project["_collections_abs"], "*.json"))):
        try:
            with open(fp, encoding="utf-8") as f:
                out.append(json.load(f))
        except Exception:
            pass
    # Auto-bind runs that carry a collection_id
    auto_bind_runs(project, out)
    # Sort by created_at desc (newest first); fallback to updated_at then id
    out.sort(key=lambda c: (c.get("created_at") or c.get("updated_at") or c.get("id") or ""), reverse=True)
    return out


def auto_bind_runs(project, collections):
    """Auto-add run_ids into collections where run JSON has `collection_id`.

    Skip if user previously removed the run (stored in `removed_run_ids`) so
    ลบแล้วไม่กลับมาโผล่อีกตอน refresh.
    """
    by_id = {c["id"]: c for c in collections}
    runs = list_runs(project)
    changed = {}
    for r in runs:
        cid = (r["_data"].get("collection_id") or "").strip()
        rid = r["_data"].get("run_id")
        if not cid or not rid or cid not in by_id:
            continue
        col = by_id[cid]
        if rid in col.get("removed_run_ids", []):
            continue  # user ลบออกไปแล้ว — ห้าม bind กลับ
        if rid not in col.get("run_ids", []):
            col.setdefault("run_ids", []).append(rid)
            changed[cid] = col
    for col in changed.values():
        save_collection(project, col)


def load_collection(project, cid):
    fp = os.path.join(project["_collections_abs"], f"{cid}.json")
    if not os.path.exists(fp):
        return None
    with open(fp, encoding="utf-8") as f:
        return json.load(f)


def save_collection(project, col):
    col["updated_at"] = datetime.now().isoformat()
    fp = os.path.join(project["_collections_abs"], f"{col['id']}.json")
    with open(fp, "w", encoding="utf-8") as f:
        json.dump(col, f, ensure_ascii=False, indent=2)
    return col


def delete_collection(project, cid):
    fp = os.path.join(project["_collections_abs"], f"{cid}.json")
    if os.path.exists(fp):
        os.remove(fp)
        return True
    return False


def load_csv_meta(project):
    """Load test metadata keyed by test_id from configured csv_dirs.

    รองรับ multi-turn CSV ที่มี column `turn` — เก็บ expected_result แยกต่อ turn ใน
    field `turns[]` เรียงตามเลข turn. Single-turn CSV ก็ยังมี expected_result ระดับ top.
    """
    meta = {}
    for rel in project.get("csv_dirs", []):
        d = os.path.join(REPO_ROOT, rel)
        if not os.path.isdir(d):
            continue
        for f in glob.glob(os.path.join(d, "*.csv")):
            with open(f, encoding="utf-8") as fh:
                for row in csv.DictReader(fh):
                    tid = row.get("test_id", "").strip()
                    if not tid:
                        continue
                    turn_raw = (row.get("turn") or "").strip()
                    expected = row.get("expected_result", "")
                    golden = row.get("golden_answer", "")
                    rubric = row.get("rubric", "")
                    if tid not in meta:
                        meta[tid] = {
                            "expected_result": expected,
                            "golden_answer": golden,
                            "rubric": rubric,
                            "turns": [],
                        }
                    # If row has a turn number, append to turns[]
                    if turn_raw:
                        try:
                            turn_num = int(turn_raw)
                            meta[tid]["turns"].append({
                                "turn": turn_num,
                                "expected_result": expected,
                                "rubric": rubric,
                                "user_text": row.get("user_text", ""),
                            })
                        except ValueError:
                            pass
    # Sort turns by turn number
    for tid in meta:
        meta[tid]["turns"].sort(key=lambda t: t.get("turn", 0))
    return meta


def _is_no_response_turn(s):
    """Return True ถ้า turn string นี้ถือเป็น no-response (empty / '(no response)' literal)."""
    t = (s or "").strip().lower()
    return t == "" or t == "(no response)"


def compute_effective(runs_data):
    """Merge multiple runs using latest-wins per test_id.

    Returns (effective_results, per_case_run_count).

    Rule: ถ้าเคสล่าสุดมี turn ใด ๆ ที่ avatar_response = empty/'(no response)' → status = REOPEN
    (infra/transport issue — ต้อง rerun ไม่ใช่ FAIL จริง)
    - Single-turn: avatar_response ทั้งก้อนว่าง = REOPEN
    - Multi-turn (split ด้วย ' → '): turn ใด turn หนึ่งว่าง/no-response = REOPEN
    """
    runs_sorted = sorted(runs_data, key=lambda r: r.get("timestamp", ""))
    latest = {}
    counts = {}
    for run in runs_sorted:
        for r in run.get("results", []):
            tid = r.get("test_id")
            if not tid:
                continue
            latest[tid] = r
            counts[tid] = counts.get(tid, 0) + 1
    effective = list(latest.values())
    # No-response rule → REOPEN (Avatar/Kiosk only — runs ที่ schema มี avatar_response)
    # Backoffice E2E results ไม่มี field นี้เลย → ห้าม flip FAIL เป็น REOPEN
    for r in effective:
        if r.get("status") != "FAIL":
            continue
        if "avatar_response" not in r:
            continue  # schema ไม่ใช่ avatar — keep FAIL ตามจริง
        resp = r.get("avatar_response") or ""
        turns = resp.split(" → ") if " → " in resp else [resp]
        if any(_is_no_response_turn(t) for t in turns):
            r["status"] = "REOPEN"
    return effective, counts


# ========== HTTP Handler ==========

class ReportHandler(BaseHTTPRequestHandler):
    projects = []

    def log_message(self, format, *args):
        # Less noisy
        sys.stderr.write(f"[{self.log_date_time_string()}] {format % args}\n")

    def _send_json(self, status, payload):
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(body)

    def _send_file(self, path, content_type):
        try:
            with open(path, "rb") as f:
                body = f.read()
        except FileNotFoundError:
            self._send_json(404, {"error": "not found"})
            return
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(body)

    def _read_json_body(self):
        length = int(self.headers.get("Content-Length") or 0)
        if length <= 0:
            return {}
        raw = self.rfile.read(length)
        try:
            return json.loads(raw)
        except Exception:
            return {}

    # ----- GET -----
    def do_GET(self):
        url = urlparse(self.path)
        path = url.path

        if path.endswith(".css"):
            return self._send_file(os.path.join(UI_DIR, os.path.basename(path)), "text/css; charset=utf-8")
        if path.endswith(".js"):
            return self._send_file(os.path.join(UI_DIR, os.path.basename(path)), "application/javascript; charset=utf-8")

        # SPA fallback — any non-/api path serves index.html
        if not path.startswith("/api/"):
            return self._send_file(os.path.join(UI_DIR, "index.html"), "text/html; charset=utf-8")

        # API
        if path == "/api/projects":
            return self._send_json(200, {
                "repo_root": REPO_ROOT,
                "projects": [
                    {"id": p["id"], "name": p["name"], "run_cmd": p.get("run_cmd", {})}
                    for p in self.projects
                ]
            })

        m = re.match(r"^/api/projects/([^/]+)/runs$", path)
        if m:
            p = find_project(m.group(1), self.projects)
            if not p:
                return self._send_json(404, {"error": "project not found"})
            runs = list_runs(p)
            # Return slim summaries (no `results`) for listing
            slim = []
            for r in runs:
                d = r["_data"]
                slim.append({
                    "run_id": d.get("run_id"),
                    "timestamp": d.get("timestamp"),
                    "preset_id": d.get("preset_id"),
                    "total": d.get("total"),
                    "passed": d.get("passed"),
                    "failed": d.get("failed"),
                    "pending": d.get("pending"),
                    "duration_seconds": d.get("duration_seconds"),
                    "phase": d.get("phase"),
                    "collection_id": d.get("collection_id"),
                })
            slim.sort(key=lambda x: x.get("timestamp", ""), reverse=True)
            return self._send_json(200, {"runs": slim})

        # Read test case rows from xlsx files (for preview table in Backoffice collection detail)
        # Query params: paths=<newline-separated paths> (URL-decoded)
        m = re.match(r"^/api/projects/([^/]+)/test-cases-content$", path)
        if m:
            p = find_project(m.group(1), self.projects)
            if not p:
                return self._send_json(404, {"error": "project not found"})
            qs = parse_qs(urlparse(self.path).query)
            paths_raw = (qs.get("paths") or [""])[0]
            paths = [s.strip() for s in re.split(r"[\n,]", paths_raw) if s.strip()]
            if not paths:
                return self._send_json(200, {"rows": []})
            try:
                import openpyxl
            except ImportError:
                return self._send_json(500, {"error": "openpyxl not installed"})
            rows = []
            for rel_path in paths:
                abs_path = os.path.join(REPO_ROOT, rel_path)
                if not os.path.isfile(abs_path):
                    continue
                try:
                    # Don't use read_only — max_row + cell access ใน read_only mode บางทีไม่เสถียร
                    wb = openpyxl.load_workbook(abs_path, data_only=True)
                    ws = wb.active
                    max_row = ws.max_row or 0
                    # Find header row — column A contains "ID" (legacy) or "TC ID" (new R2/R3 template)
                    header_row = None
                    id_aliases = ("ID", "TC ID")
                    for r in range(1, min(max_row + 1, 20)):
                        v = ws.cell(r, 1).value
                        if v and str(v).strip() in id_aliases:
                            header_row = r
                            break
                    if not header_row:
                        continue
                    # Auto-detect columns from header — supports legacy and new R2/R3 templates
                    header_map = {}
                    for col_idx in range(1, ws.max_column + 1):
                        h = ws.cell(header_row, col_idx).value
                        if h:
                            header_map[str(h).strip()] = col_idx
                    # Helper: first matching column name wins
                    def _pick(*names):
                        for n in names:
                            if n in header_map:
                                return header_map[n]
                        return None
                    # In new R2/R3 template: "Step #" = step number, "Step" = step content (test step).
                    # In legacy template: "Step" = step number, "Test Step" = step content.
                    if "Step #" in header_map:
                        col_step      = header_map.get("Step #")
                        col_test_step = header_map.get("Step")
                    else:
                        col_step      = header_map.get("Step")
                        col_test_step = header_map.get("Test Step")
                    col_expected  = _pick("Expected Result", "Expect Result")
                    col_scenario  = _pick("Scenario", "ชื่อ Test Case")
                    col_precond   = _pick("Pre-conditions", "Pre-condition")
                    col_test_data = _pick("Test Data", "Permissions ที่ติ๊ก")
                    # Type detection — two schemas:
                    #   Legacy: single "Automation" column with values like "[auto]", "🤖 Script", "[manual]"
                    #   R2/R3:  separate "Script" and "Manual" columns (each ticked with ✅ if applicable)
                    col_automation = header_map.get("Automation")  # legacy single col
                    col_type_script = header_map.get("Script")
                    col_type_manual = header_map.get("Manual")
                    if not (col_test_step and col_expected):
                        continue

                    last_id = ""
                    last_scenario = ""
                    last_precond = ""
                    last_test_data = ""
                    last_type = ""

                    def _split_sections(text):
                        """Split Test Step text by '### Step N — title' divider into sections.
                        Each section contains numbered items (1. 2. 3...).
                        Returns [{title: str, steps: [(num, content), ...]}, ...]
                        If no '### Step' divider found, returns [] (caller falls back to flat split)."""
                        if not text:
                            return []
                        import re as _re
                        # Section divider: "### Step 1 — title", "### Step 2 — title", etc.
                        section_pat = _re.compile(r"^\s*#{2,4}\s+Step\s+\d+(?:\s*[—\-:]\s*(.*))?$", _re.IGNORECASE)
                        lines = str(text).split("\n")
                        # Check if there's any section divider — if not, return [] for fallback
                        if not any(section_pat.match(ln) for ln in lines):
                            return []
                        sections = []
                        cur_section = None
                        cur_lines = []
                        for ln in lines:
                            m = section_pat.match(ln)
                            if m:
                                # Close current section
                                if cur_section is not None:
                                    cur_section["steps"] = _split_numbered("\n".join(cur_lines))
                                    sections.append(cur_section)
                                cur_section = {"title": (m.group(1) or "").strip(), "steps": []}
                                cur_lines = []
                            else:
                                cur_lines.append(ln)
                        # Close final section
                        if cur_section is not None:
                            cur_section["steps"] = _split_numbered("\n".join(cur_lines))
                            sections.append(cur_section)
                        # Drop sections that have no steps
                        return [s for s in sections if s["steps"]]

                    def _split_sections_expected(text):
                        """Split Expected Result text by '**Step N — title:**' bold header.
                        Each section gets its body text (bullets/paragraphs that follow).
                        Returns [{title: str, body: str}, ...] aligned with Test Step sections."""
                        if not text:
                            return []
                        import re as _re
                        # Expected divider: "**Step 1 — title:**" or "**Step 1:**"
                        header_pat = _re.compile(r"^\s*\*\*Step\s+\d+(?:\s*[—\-:]\s*(.*?))?:\s*\*\*\s*$", _re.IGNORECASE)
                        lines = str(text).split("\n")
                        if not any(header_pat.match(ln) for ln in lines):
                            return []
                        sections = []
                        cur = None
                        for ln in lines:
                            m = header_pat.match(ln)
                            if m:
                                if cur is not None:
                                    cur["body"] = cur["body"].strip()
                                    sections.append(cur)
                                cur = {"title": (m.group(1) or "").strip().rstrip(":").strip(), "body": ""}
                            elif cur is not None:
                                cur["body"] += ln + "\n"
                        if cur is not None:
                            cur["body"] = cur["body"].strip()
                            sections.append(cur)
                        return sections

                    def _split_numbered(text, renumber=True):
                        """Split a 'Test Step' or 'Expected Result' blob containing numbered items.
                        Supports formats:
                          - '1. text...' (Test Step)
                          - '- Step 1: text...' (Expected Result, bullet form)
                          - 'Step 1: text...' (Expected Result, plain form)
                        Skipped (treated as dividers, not attached to previous step):
                          - Role headers '**Role (context):**'
                          - Markdown section headers '### Step N — Title' / '## ...'
                          - Horizontal dividers '---'
                        If renumber=True (default — for Test Step): step numbers are renumbered
                        sequentially (1, 2, 3...) so restarts in new sections become the next number.
                        If renumber=False (for Expected Result): preserve original numbers verbatim
                        so 'Step 5: x' maps to test step 5.
                        Returns list of (step_num, content). [] if no markers found."""
                        if not text:
                            return []
                        import re as _re
                        # Match "N. xxx" or "- Step N: xxx" or "Step N: xxx" (but NOT "### Step N")
                        item_pat = _re.compile(r"^\s*(?:-\s*)?(?:Step\s+)?(\d+)[.:]\s+(.*)$", _re.IGNORECASE)
                        # Headers/dividers to skip
                        # Bold role header: "**Super Admin (context):**", "**Role:**"
                        role_bold_pat = _re.compile(r"^\s*\*\*[^*]+:\s*\*\*\s*$", _re.IGNORECASE)
                        # Plain role header (ends with ":" but is not a step item):
                        # "Super Admin (admin context):", "permissions@acme.test (test user context):"
                        # Heuristic: ends with ":" and contains "(...)" — typical role context marker
                        role_plain_pat = _re.compile(r"^[^0-9].*\(.*\):\s*$")
                        heading_pat = _re.compile(r"^\s*#{1,6}\s+", )
                        divider_pat = _re.compile(r"^\s*-{3,}\s*$")
                        bold_step_header = _re.compile(r"^\s*\*\*Step\s+\d+.*\*\*\s*$", _re.IGNORECASE)
                        lines = str(text).split("\n")
                        items = []
                        for ln in lines:
                            stripped = ln.strip()
                            if not stripped:
                                if items:
                                    items[-1][1] += "\n"
                                continue
                            if (heading_pat.match(ln) or divider_pat.match(ln)
                                    or role_bold_pat.match(ln) or role_plain_pat.match(ln)
                                    or bold_step_header.match(ln)):
                                continue
                            m = item_pat.match(ln)
                            if m:
                                snum = str(len(items) + 1) if renumber else m.group(1)
                                items.append([snum, m.group(2)])
                            elif items:
                                items[-1][1] += "\n" + ln
                        return [(n, c.strip()) for n, c in items]

                    for r in range(header_row + 1, max_row + 1):
                        tid = ws.cell(r, 1).value
                        step_raw = ws.cell(r, col_step).value if col_step else None
                        test_step_raw = ws.cell(r, col_test_step).value
                        expected_raw = ws.cell(r, col_expected).value
                        if not any([tid, step_raw, test_step_raw, expected_raw]):
                            continue
                        if tid:
                            last_id = str(tid).strip()
                            last_scenario = str(ws.cell(r, col_scenario).value or "").strip() if col_scenario else ""
                            last_precond = str(ws.cell(r, col_precond).value or "").strip() if col_precond else ""
                            last_test_data = str(ws.cell(r, col_test_data).value or "").strip() if col_test_data else ""
                            # 1) Legacy single "Automation" column
                            auto_raw = str(ws.cell(r, col_automation).value or "").strip() if col_automation else ""
                            auto_lower = auto_raw.lower()
                            if "[auto]" in auto_lower or "🤖" in auto_raw:
                                last_type = "Script"
                            elif "[manual]" in auto_lower or "🧑" in auto_raw or "👤" in auto_raw:
                                last_type = "Manual"
                            else:
                                last_type = ""

                            # 2) R2/R3 two-column schema (Script / Manual) — any ticked = applicable.
                            # Treat any non-empty mark in these columns as a tick (✅, ✓, x, yes, true).
                            def _is_ticked(col_idx):
                                if not col_idx:
                                    return False
                                v = str(ws.cell(r, col_idx).value or "").strip().lower()
                                return v and v not in ("∅", "-", "none", "no", "false", "0")
                            if not last_type:
                                # Prefer Script over Manual when both are ticked.
                                if _is_ticked(col_type_script):
                                    last_type = "Script"
                                elif _is_ticked(col_type_manual):
                                    last_type = "Manual"

                        test_step_text = str(test_step_raw or "").strip()
                        expected_text = str(expected_raw or "").strip()

                        # NEW format: Test Step contains "1. ...\n2. ...\n" → split into per-step rows
                        if col_step is None:
                            # Try section-aware split first (handles "### Step N — title" dividers)
                            step_sections = _split_sections(test_step_text)
                            exp_sections = _split_sections_expected(expected_text)

                            if step_sections and any(len(sec["steps"]) for sec in step_sections):
                                # Flatten sections into a single sequential numbering
                                flat_step_rows = []  # list of (step_num, content, section_idx)
                                for sec_idx, sec in enumerate(step_sections):
                                    for (_, scontent) in sec["steps"]:
                                        flat_step_rows.append((str(len(flat_step_rows) + 1), scontent, sec_idx))

                                # Build expected map: section_idx → expected text (full block per section)
                                exp_by_section = {idx: sec["body"] for idx, sec in enumerate(exp_sections)}
                                # If expected has no sections at all but has body, attach all to section 0
                                if not exp_sections and expected_text:
                                    exp_by_section[0] = expected_text

                                # Emit one row per step. Expected goes to LAST step of each section
                                # (steps within a section are actions; expected = the verifiable
                                # outcome after all actions in that section complete).
                                last_of_section = {}
                                for snum, scontent, sec_idx in flat_step_rows:
                                    last_of_section[sec_idx] = snum  # overwrites — last write wins

                                fallback_exp_used = False
                                for i, (snum, scontent, sec_idx) in enumerate(flat_step_rows):
                                    exp = ""
                                    if last_of_section.get(sec_idx) == snum:
                                        exp = exp_by_section.get(sec_idx, "")
                                        # If still empty AND no expected attached anywhere yet, use full text
                                        if not exp and not fallback_exp_used and not exp_by_section:
                                            exp = expected_text
                                            fallback_exp_used = True
                                    rows.append({
                                        "id": last_id,
                                        "type": last_type,
                                        "scenario": last_scenario if i == 0 else "",
                                        "pre_conditions": last_precond if i == 0 else "",
                                        "test_data": last_test_data if i == 0 else "",
                                        "step": snum,
                                        "test_step": scontent,
                                        "expected_result": exp,
                                        "source_file": rel_path,
                                    })
                            else:
                                # Fallback to flat split (no sections)
                                step_items = _split_numbered(test_step_text, renumber=True)
                                # Expected uses ORIGINAL step numbers (so "Step 5: x" maps to test step 5)
                                expected_items = _split_numbered(expected_text, renumber=False)
                                exp_map = {n: c for n, c in expected_items}
                                if step_items:
                                    fallback_exp = expected_text if not expected_items else ""
                                    for i, (snum, scontent) in enumerate(step_items):
                                        exp = exp_map.get(snum, "")
                                        if i == 0 and not exp:
                                            exp = fallback_exp
                                        rows.append({
                                            "id": last_id,
                                            "type": last_type,
                                            "scenario": last_scenario if i == 0 else "",
                                            "pre_conditions": last_precond if i == 0 else "",
                                            "test_data": last_test_data if i == 0 else "",
                                            "step": snum,
                                            "test_step": scontent,
                                            "expected_result": exp,
                                            "source_file": rel_path,
                                        })
                                else:
                                    # No numbered list at all — render as a single row
                                    rows.append({
                                        "id": last_id,
                                        "type": last_type,
                                        "scenario": last_scenario,
                                        "pre_conditions": last_precond,
                                        "test_data": last_test_data,
                                        "step": "1",
                                        "test_step": test_step_text,
                                        "expected_result": expected_text,
                                        "source_file": rel_path,
                                    })
                        else:
                            # OLD format: Step column already has the step number, 1 row per step
                            if step_raw is None:
                                step_str = ""
                            elif isinstance(step_raw, float) and step_raw.is_integer():
                                step_str = str(int(step_raw))
                            else:
                                step_str = str(step_raw).strip()
                            rows.append({
                                "id": last_id,
                                "type": last_type,
                                "scenario": last_scenario,
                                "pre_conditions": last_precond,
                                "test_data": last_test_data,
                                "step": step_str,
                                "test_step": test_step_text,
                                "expected_result": expected_text,
                                "source_file": rel_path,
                            })
                except Exception as e:
                    import traceback
                    rows.append({"id": "ERROR", "scenario": f"Cannot read {rel_path}", "error": str(e) + " | " + traceback.format_exc()[:300]})
            return self._send_json(200, {"rows": rows})

        # List test case xlsx files in project csv_dirs (for dropdown in Create Collection modal)
        m = re.match(r"^/api/projects/([^/]+)/test-cases$", path)
        if m:
            p = find_project(m.group(1), self.projects)
            if not p:
                return self._send_json(404, {"error": "project not found"})
            files = []
            seen_paths = set()  # dedupe across overlapping csv_dirs
            for d in p.get("csv_dirs", []):
                abs_dir = os.path.join(REPO_ROOT, d)
                if not os.path.isdir(abs_dir):
                    continue
                # Recursive walk — pick up xlsx in any subfolder (no manual config per subdir)
                for root, _, filenames in os.walk(abs_dir):
                    for fname in sorted(filenames):
                        if not fname.endswith(".xlsx") or fname.startswith("~$"):
                            continue
                        abs_path = os.path.join(root, fname)
                        rel_path = os.path.relpath(abs_path, REPO_ROOT)
                        if rel_path in seen_paths:
                            continue
                        seen_paths.add(rel_path)
                        rel_folder = os.path.relpath(root, REPO_ROOT)
                        files.append({
                            "path": rel_path,
                            "name": fname,
                            "folder": rel_folder,
                        })
            # Stable order: by path so UI grouping is consistent
            files.sort(key=lambda f: f["path"])
            return self._send_json(200, {"test_cases": files})

        m = re.match(r"^/api/projects/([^/]+)/runs/([^/]+)$", path)
        if m:
            p = find_project(m.group(1), self.projects)
            if not p:
                return self._send_json(404, {"error": "project not found"})
            r = load_run(p, m.group(2))
            if not r:
                return self._send_json(404, {"error": "run not found"})
            return self._send_json(200, {"run": r["_data"], "meta": load_csv_meta(p)})

        m = re.match(r"^/api/projects/([^/]+)/collections$", path)
        if m:
            p = find_project(m.group(1), self.projects)
            if not p:
                return self._send_json(404, {"error": "project not found"})
            cols = list_collections(p)
            # Compute summary (total/passed/failed/pending/reopened) per collection from effective merge.
            # Apply per-case user verdict override from collection.case_statuses — overrides win.
            for c in cols:
                run_datas = []
                for rid in c.get("run_ids", []):
                    r = load_run(p, rid)
                    if r:
                        run_datas.append(r["_data"])
                effective, _ = compute_effective(run_datas)
                overrides = c.get("case_statuses") or {}
                for x in effective:
                    tid = x.get("test_id")
                    if tid and tid in overrides:
                        x["status"] = overrides[tid]
                total = len(effective)
                passed = sum(1 for x in effective if x.get("status") == "PASS")
                failed = sum(1 for x in effective if x.get("status") == "FAIL")
                pending = sum(1 for x in effective if x.get("status") == "PENDING")
                reopened = sum(1 for x in effective if x.get("status") == "REOPEN")
                improvement = sum(1 for x in effective if x.get("status") == "IMPROVEMENT")
                warn = sum(1 for x in effective if x.get("status") == "WARN")
                c["summary"] = {
                    "total": total,
                    "passed": passed,
                    "failed": failed,
                    "pending": pending,
                    "reopened": reopened,
                    "improvement": improvement,
                    "warn": warn,
                    "pass_rate": round(passed / total * 100, 1) if total else 0,
                }
            return self._send_json(200, {"collections": cols})

        m = re.match(r"^/api/projects/([^/]+)/jira-all$", path)
        if m:
            p = find_project(m.group(1), self.projects)
            if not p:
                return self._send_json(404, {"error": "project not found"})
            if not JIRA_CONFIG:
                return self._send_json(200, {"configured": False, "collections": []})
            fresh = parse_qs(url.query).get("fresh", ["0"])[0] == "1"
            cols = list_collections(p)
            out = []
            for c in cols:
                cid = c.get("id")
                if not cid:
                    continue
                issues = fetch_jira_issues_for_collection(cid, fresh=fresh) or []
                if not issues:
                    continue  # skip collections with no issues — keep output compact
                out.append({
                    "collection_id": cid,
                    "collection_name": c.get("name", cid),
                    "collection_type": c.get("type", ""),
                    "created_at": c.get("created_at", ""),
                    "issues": issues,
                })
            # sort by most recent collection first
            out.sort(key=lambda x: x.get("created_at", ""), reverse=True)
            return self._send_json(200, {
                "configured": True,
                "base_url": JIRA_CONFIG["base_url"],
                "project_key": JIRA_CONFIG["project_key"],
                "collections": out,
            })

        m = re.match(r"^/api/projects/([^/]+)/collections/([^/]+)/jira-issues$", path)
        if m:
            p = find_project(m.group(1), self.projects)
            if not p:
                return self._send_json(404, {"error": "project not found"})
            cid = m.group(2)
            if not ID_SAFE_RE.match(cid):
                return self._send_json(400, {"error": "invalid collection id"})
            fresh = parse_qs(url.query).get("fresh", ["0"])[0] == "1"
            issues = fetch_jira_issues_for_collection(cid, fresh=fresh)
            if issues is None:
                return self._send_json(200, {"configured": False, "issues": []})
            return self._send_json(200, {
                "configured": True,
                "base_url": JIRA_CONFIG["base_url"],
                "project_key": JIRA_CONFIG["project_key"],
                "issues": issues,
            })

        m = re.match(r"^/api/projects/([^/]+)/collections/([^/]+)$", path)
        if m:
            p = find_project(m.group(1), self.projects)
            if not p:
                return self._send_json(404, {"error": "project not found"})
            col = load_collection(p, m.group(2))
            if not col:
                return self._send_json(404, {"error": "collection not found"})
            # Auto-bind new runs that carry this collection_id (e.g. fresh `npm run e2e`
            # results that haven't been visited via the collection list yet).
            auto_bind_runs(p, [col])
            # Re-read in case auto_bind_runs wrote new run_ids back to disk
            col = load_collection(p, m.group(2)) or col
            # Hydrate: include full run data + effective merge
            run_datas = []
            cwd_abs = os.path.join(REPO_ROOT, (p.get("run_cmd") or {}).get("cwd", ""))
            for rid in col.get("run_ids", []):
                r = load_run(p, rid)
                if r:
                    d = dict(r["_data"])  # shallow copy เพื่อไม่กระทบ cache
                    abs_fp = r["_file"]
                    # path สำหรับ resync.py (relative ต่อ project cwd)
                    try:
                        d["_file_rel"] = os.path.relpath(abs_fp, cwd_abs)
                    except Exception:
                        d["_file_rel"] = abs_fp
                    run_datas.append(d)
            effective, counts = compute_effective(run_datas)
            total = len(effective)
            passed = sum(1 for x in effective if x.get("status") == "PASS")
            failed = sum(1 for x in effective if x.get("status") == "FAIL")
            pending = sum(1 for x in effective if x.get("status") == "PENDING")
            reopened = sum(1 for x in effective if x.get("status") == "REOPEN")
            improvement = sum(1 for x in effective if x.get("status") == "IMPROVEMENT")
            warn = sum(1 for x in effective if x.get("status") == "WARN")
            return self._send_json(200, {
                "collection": col,
                "runs": run_datas,
                "effective": {
                    "results": effective,
                    "total": total,
                    "passed": passed,
                    "failed": failed,
                    "pending": pending,
                    "reopened": reopened,
                    "improvement": improvement,
                    "warn": warn,
                    "pass_rate": round(passed / total * 100, 1) if total else 0,
                    "per_case_run_count": counts,
                },
                "meta": load_csv_meta(p),
            })

        return self._send_json(404, {"error": "not found", "path": path})

    # ----- POST -----
    def do_POST(self):
        path = urlparse(self.path).path

        m = re.match(r"^/api/projects/([^/]+)/collections$", path)
        if m:
            p = find_project(m.group(1), self.projects)
            if not p:
                return self._send_json(404, {"error": "project not found"})
            body = self._read_json_body()
            name = (body.get("name") or "").strip()
            if not name:
                return self._send_json(400, {"error": "name required"})
            cid = body.get("id") or f"col-{datetime.now().strftime('%Y%m%d-%H%M%S')}-{uuid.uuid4().hex[:6]}"
            if not ID_SAFE_RE.match(cid):
                return self._send_json(400, {"error": "invalid id"})
            col = {
                "id": cid,
                "name": name,
                "description": body.get("description", ""),
                "type": body.get("type", ""),
                "category": body.get("category", ""),
                "collection_url": body.get("collection_url", ""),
                "preset_url": body.get("preset_url", ""),
                "preset": body.get("preset", ""),
                "excel_path": body.get("excel_path", ""),
                "test_case_path": body.get("test_case_path", ""),
                "result_url": body.get("result_url", ""),
                "jira_epic": body.get("jira_epic", ""),
                "jira_run": body.get("jira_run", ""),
                "kb_path": body.get("kb_path", ""),
                "created_at": datetime.now().isoformat(),
                "updated_at": datetime.now().isoformat(),
                "run_ids": body.get("run_ids", []),
            }
            save_collection(p, col)
            return self._send_json(201, {"collection": col})

        m = re.match(r"^/api/projects/([^/]+)/collections/([^/]+)/runs$", path)
        if m:
            p = find_project(m.group(1), self.projects)
            if not p:
                return self._send_json(404, {"error": "project not found"})
            col = load_collection(p, m.group(2))
            if not col:
                return self._send_json(404, {"error": "collection not found"})
            body = self._read_json_body()
            rid = body.get("run_id")
            if not rid:
                return self._send_json(400, {"error": "run_id required"})
            # ถ้า user เคยลบแล้วเพิ่มกลับเอง — เอาออกจาก blacklist ให้ auto_bind ทำงานได้อีก
            removed = col.get("removed_run_ids", [])
            if rid in removed:
                removed.remove(rid)
            if rid not in col["run_ids"]:
                col["run_ids"].append(rid)
            save_collection(p, col)
            return self._send_json(200, {"collection": col})

        m = re.match(r"^/api/projects/([^/]+)/collections/([^/]+)/screenshot$", path)
        if m:
            p = find_project(m.group(1), self.projects)
            if not p:
                return self._send_json(404, {"error": "project not found"})
            cid = m.group(2)
            if not ID_SAFE_RE.match(cid):
                return self._send_json(400, {"error": "invalid collection id"})
            col = load_collection(p, cid)
            if not col:
                return self._send_json(404, {"error": "collection not found"})
            # Screenshot script (lives in tests/kiosk-avatar/shared/) is reused
            # for other projects via --base-url + --out-root flags.
            script_dir = os.path.join(REPO_ROOT, "tests", "kiosk-avatar")
            script_path = os.path.join(script_dir, "shared", "capture_screenshots.py")
            venv_py = os.path.join(script_dir, ".venv", "bin", "python")
            if not os.path.isfile(script_path):
                return self._send_json(500, {"error": f"script not found: {script_path}"})
            py = venv_py if os.path.isfile(venv_py) else sys.executable

            # Determine port from request Host header (defaults 8088)
            host_header = self.headers.get("Host") or "localhost:8088"
            base_url = f"http://{host_header}/{p['id']}/collections"

            # Output root — uses project's results_dir (or kiosk-avatar's default if missing)
            out_rel_root = p.get("results_dir") or "tests/kiosk-avatar/results"

            try:
                proc = subprocess.run(
                    [py, "shared/capture_screenshots.py", cid,
                     "--base-url", base_url,
                     "--out-root", os.path.join(REPO_ROOT, out_rel_root)],
                    cwd=script_dir,
                    capture_output=True,
                    text=True,
                    timeout=120,
                )
            except subprocess.TimeoutExpired:
                return self._send_json(504, {"error": "screenshot timed out (120s)"})
            now = datetime.now()
            out_rel = f"{out_rel_root}/{now.year:04d}/{now.month:02d}/{now.day:02d}/screenshot/{cid}"
            return self._send_json(200 if proc.returncode == 0 else 500, {
                "ok": proc.returncode == 0,
                "stdout": proc.stdout[-2000:],
                "stderr": proc.stderr[-1000:],
                "output_dir": out_rel,
            })

        m = re.match(r"^/api/projects/([^/]+)/collections/([^/]+)/pdf$", path)
        if m:
            p = find_project(m.group(1), self.projects)
            if not p:
                return self._send_json(404, {"error": "project not found"})
            cid = m.group(2)
            if not ID_SAFE_RE.match(cid):
                return self._send_json(400, {"error": "invalid collection id"})
            col = load_collection(p, cid)
            if not col:
                return self._send_json(404, {"error": "collection not found"})
            # PDF script ใช้ chrome headless --print-to-pdf — อยู่ที่เดียวกับ screenshot script
            script_dir = os.path.join(REPO_ROOT, "tests", "kiosk-avatar")
            script_path = os.path.join(script_dir, "shared", "capture_pdf.py")
            venv_py = os.path.join(script_dir, ".venv", "bin", "python")
            if not os.path.isfile(script_path):
                return self._send_json(500, {"error": f"script not found: {script_path}"})
            py = venv_py if os.path.isfile(venv_py) else sys.executable

            host_header = self.headers.get("Host") or "localhost:8088"
            base_url = f"http://{host_header}/{p['id']}/collections"
            out_rel_root = p.get("results_dir") or "tests/kiosk-avatar/results"

            try:
                proc = subprocess.run(
                    [py, "shared/capture_pdf.py", cid,
                     "--base-url", base_url,
                     "--out-root", os.path.join(REPO_ROOT, out_rel_root)],
                    cwd=script_dir,
                    capture_output=True,
                    text=True,
                    timeout=120,
                )
            except subprocess.TimeoutExpired:
                return self._send_json(504, {"error": "PDF capture timed out (120s)"})
            now = datetime.now()
            out_rel = f"{out_rel_root}/{now.year:04d}/{now.month:02d}/{now.day:02d}/pdf/{cid}.pdf"
            return self._send_json(200 if proc.returncode == 0 else 500, {
                "ok": proc.returncode == 0,
                "stdout": proc.stdout[-2000:],
                "stderr": proc.stderr[-1000:],
                "output_path": out_rel,
            })

        return self._send_json(404, {"error": "not found"})

    # ----- PATCH -----
    def do_PATCH(self):
        path = urlparse(self.path).path
        m = re.match(r"^/api/projects/([^/]+)/collections/([^/]+)$", path)
        if m:
            p = find_project(m.group(1), self.projects)
            if not p:
                return self._send_json(404, {"error": "project not found"})
            col = load_collection(p, m.group(2))
            if not col:
                return self._send_json(404, {"error": "collection not found"})
            body = self._read_json_body()
            for k in ("name", "description", "type", "category", "collection_url", "preset_url", "preset",
                     "excel_path", "test_case_path", "result_url", "jira_epic", "jira_run", "kb_path",
                     "case_statuses", "case_remarks"):
                if k in body:
                    col[k] = body[k]
            # Special: case_status_update = single status patch (merge into case_statuses dict)
            if "case_status_update" in body:
                upd = body["case_status_update"]  # {"case_id": "BO-LGN-E2E-001", "status": "PASS"}
                if upd and upd.get("case_id"):
                    if "case_statuses" not in col or not isinstance(col["case_statuses"], dict):
                        col["case_statuses"] = {}
                    col["case_statuses"][upd["case_id"]] = upd.get("status", "PENDING")
            # Special: case_remark_update = single remark patch (merge into case_remarks dict)
            if "case_remark_update" in body:
                upd = body["case_remark_update"]  # {"case_id": "BO-LGN-E2E-001", "remark": "text"}
                if upd and upd.get("case_id"):
                    if "case_remarks" not in col or not isinstance(col["case_remarks"], dict):
                        col["case_remarks"] = {}
                    col["case_remarks"][upd["case_id"]] = upd.get("remark", "")
            save_collection(p, col)
            return self._send_json(200, {"collection": col})
        return self._send_json(404, {"error": "not found"})

    # ----- DELETE -----
    def do_DELETE(self):
        path = urlparse(self.path).path

        m = re.match(r"^/api/projects/([^/]+)/collections/([^/]+)$", path)
        if m:
            p = find_project(m.group(1), self.projects)
            if not p:
                return self._send_json(404, {"error": "project not found"})
            ok = delete_collection(p, m.group(2))
            return self._send_json(200 if ok else 404, {"ok": ok})

        m = re.match(r"^/api/projects/([^/]+)/collections/([^/]+)/runs/([^/]+)$", path)
        if m:
            p = find_project(m.group(1), self.projects)
            if not p:
                return self._send_json(404, {"error": "project not found"})
            col = load_collection(p, m.group(2))
            if not col:
                return self._send_json(404, {"error": "collection not found"})
            rid = m.group(3)
            if rid in col["run_ids"]:
                col["run_ids"].remove(rid)
            # Record ว่า user ลบ rid นี้ เพื่อกัน auto_bind ไม่ให้เอากลับมา
            removed = col.setdefault("removed_run_ids", [])
            if rid not in removed:
                removed.append(rid)
            save_collection(p, col)
            return self._send_json(200, {"collection": col})

        m = re.match(r"^/api/projects/([^/]+)/runs/([^/]+)$", path)
        if m:
            p = find_project(m.group(1), self.projects)
            if not p:
                return self._send_json(404, {"error": "project not found"})
            rid = m.group(2)
            r = load_run(p, rid)
            if not r:
                return self._send_json(404, {"error": "run not found"})
            try:
                os.remove(r["_file"])
            except OSError as e:
                return self._send_json(500, {"error": f"failed to delete: {e}"})
            # Cleanup: remove run from any collection that contains it
            for col in list_collections(p):
                if rid in col.get("run_ids", []):
                    col["run_ids"].remove(rid)
                    save_collection(p, col)
            return self._send_json(200, {"ok": True, "deleted": rid})

        return self._send_json(404, {"error": "not found"})


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--port", type=int, default=8088)
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument(
        "--repo-root",
        default=None,
        help="path ของ main repo (realfact) ที่เก็บ test data; "
             "ถ้าไม่ระบุ จะอ่านจาก env REALFACT_REPO_ROOT แล้ว auto-detect",
    )
    args = parser.parse_args()

    global REPO_ROOT, JIRA_CONFIG_PATH, JIRA_CONFIG
    REPO_ROOT = resolve_repo_root(args.repo_root)
    JIRA_CONFIG_PATH = os.path.join(REPO_ROOT, ".claude", "jira_config.md")

    if not os.path.isdir(REPO_ROOT):
        print(f"[report-server] ERROR: ไม่พบ main repo ที่ '{REPO_ROOT}'")
        print("[report-server] ระบุด้วย --repo-root /path/to/realfact "
              "หรือ env REALFACT_REPO_ROOT")
        sys.exit(1)
    if not os.path.isdir(os.path.join(REPO_ROOT, "tests", "kiosk-avatar")):
        print(f"[report-server] WARNING: '{REPO_ROOT}' ไม่มี tests/kiosk-avatar "
              "— อาจชี้ผิด repo (data จะว่างเปล่า)")

    ReportHandler.projects = load_projects()
    JIRA_CONFIG = load_jira_config()

    print(f"[report-server] Repo root: {REPO_ROOT}")
    print(f"[report-server] Loaded {len(ReportHandler.projects)} projects:")
    for p in ReportHandler.projects:
        print(f"  - {p['id']} ({p['name']})  →  {p['results_dir']}")
    if JIRA_CONFIG:
        print(f"[report-server] JIRA: enabled (project {JIRA_CONFIG['project_key']} on {JIRA_CONFIG['base_url']})")
    else:
        print(f"[report-server] JIRA: disabled (no {JIRA_CONFIG_PATH})")
    print(f"[report-server] Listening at http://{args.host}:{args.port}")
    print(f"[report-server] (Ctrl+C to stop)")

    httpd = ThreadingHTTPServer((args.host, args.port), ReportHandler)
    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        print("\n[report-server] shutting down")
    finally:
        httpd.server_close()


if __name__ == "__main__":
    main()
